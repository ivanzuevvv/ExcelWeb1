import os
import argparse
import openpyxl
from copy import copy
import logging
import plugins.base

class pluginClass(plugins.base.basePlugin):
    def __init__(self):
        pass
    def run(self, pluginInput, pluginOutput):
        py_logger = logging.getLogger(__name__)
        py_logger.setLevel(logging.INFO)
        py_handler = logging.FileHandler(f"{pluginOutput}/{__name__}.log", mode='w')
        py_formatter = logging.Formatter("%(name)s %(asctime)s %(levelname)s %(message)s")
        py_handler.setFormatter(py_formatter)
        py_logger.addHandler(py_handler)

        def makeReport(wbSources, wbSampleFileName, wbReportFileName):
            """
            Сформировать отчет.
            """          
            def copy_cell(src_sheet, src_row, src_col, 
                            tgt_sheet, tgt_row, tgt_col,
                            copy_style=True, copy_style_only=False):
                """
                Копировать содержимое ячейки.
                """
                cell = src_sheet.cell(src_row, src_col)
                if not copy_style_only:                
                    new_cell = tgt_sheet.cell(tgt_row, tgt_col, cell.value)
                else:
                    new_cell = tgt_sheet.cell(tgt_row, tgt_col)
                    # копируем еще и скорректированные формулы (замена номера строки) 
                    if cell.data_type == 'f':
                        new_cell.value = str(cell.value).replace(str(src_row), str(tgt_row))
                
                if cell.has_style and copy_style:
                    # print(dir(new_cell))
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
                    # new_cell._style = copy(cell._style)

            wbSourceCount = len(wbSources)
            # print("wbSourceCount=", wbSourceCount)

            if wbSourceCount > 0:
                wbSample = openpyxl.load_workbook(wbSampleFileName)    # открываем образец
                wbReport = openpyxl.Workbook()
                wsReport01 = wbReport.active
                wsReport02 = wbReport.create_sheet(wbSample.worksheets[1].title)            
                
                p1y1 = list()
                p1y2 = list()
                p2y1 = list()
                p2y2 = list()

                sp1y1 = list()
                sp1y2 = list()
                sp2y1 = list()
                sp2y2 = list()

                wsReport01.sheet_format = copy(wbSample.worksheets[0].sheet_format)
                wsReport02.sheet_format = copy(wbSample.worksheets[1].sheet_format)
                wsReport01.title = wbSample.worksheets[0].title
                    
                # ширина столбцов лист 1
                for iColumn in range(wbSample.worksheets[0].max_column):
                    iColumn_letter = wsReport01.cell(1,iColumn+1).column_letter
                    if iColumn_letter in wbSample.worksheets[0].column_dimensions:
                        iWidth = wbSample.worksheets[0].column_dimensions[iColumn_letter].width
                    wsReport01.column_dimensions[iColumn_letter].width = iWidth
                # ширина столбцов лист 2
                for iColumn in range(wbSample.worksheets[1].max_column):
                    iColumn_letter = wsReport02.cell(1,iColumn+1).column_letter
                    if iColumn_letter in wbSample.worksheets[1].column_dimensions:
                        iWidth = wbSample.worksheets[1].column_dimensions[iColumn_letter].width
                    wsReport02.column_dimensions[iColumn_letter].width = iWidth
                # for k, cd in wbSample.worksheets[0].column_dimensions.items():
                #     wsReport01.column_dimensions[k].width = cd.width

                # высота строк лист 1
                for k, rd in wbSample.worksheets[0].row_dimensions.items():
                    wsReport01.row_dimensions[k].height = rd.height
                # объединенные ячейки лист 1
                wsReport01.merged_cells = copy(wbSample.worksheets[0].merged_cells)
                # высота строк лист 2
                for k, rd in wbSample.worksheets[1].row_dimensions.items():
                    wsReport02.row_dimensions[k].height = rd.height
                # объединенные ячейки лист 2
                wsReport02.merged_cells = copy(wbSample.worksheets[1].merged_cells)

                # копируем строки с 1 по 7 из образца лист 1
                for iRow in wbSample.worksheets[0].iter_rows(min_row=1, max_col=wbSample.worksheets[0].max_column, max_row=7):
                    for iCell in iRow:
                        copy_cell(wbSample.worksheets[0], iCell.row, iCell.column,
                                wsReport01, iCell.row, iCell.column)            
                # копируем строки с 1 по 7 из образца лист 2
                for iRow in wbSample.worksheets[1].iter_rows(min_row=1, max_col=wbSample.worksheets[0].max_column, max_row=7):
                    for iCell in iRow:
                        copy_cell(wbSample.worksheets[1], iCell.row, iCell.column,
                                wsReport02, iCell.row, iCell.column)            
                p1y1.clear()
                p1y2.clear()
                p2y1.clear()
                p2y2.clear()
                sp1y1.clear()
                sp1y2.clear()
                sp2y1.clear()
                sp2y2.clear()
            ###########################################
                for iSourceFileName in wbSources:
                    try:
                        py_logger.info(f"Обработка файла {iSourceFileName}")                        
                        wbSource = openpyxl.load_workbook(iSourceFileName, data_only=True)    # открываем файл с данными
                        wsSource = wbSource.worksheets[0]                    
                        rows = wsSource.iter_rows(min_row=1, max_col=wsSource.max_column, max_row=wsSource.max_row)
                        #   определяем первые и последние строки диапазонов
                        for row in rows:                        
                            if row[1].value == "Услуга 1. Услуги общежитий для СУ и ЮУ МУО ПАО \"Газпром\"":
                                p1y1FirstRow = row[1].row
                            if row[1].value == "Услуга 2. Услуги спортивных учреждений для СУ и ЮУ МУО ПАО \"Газпром\"":
                                p1y2FirstRow = row[1].row
                            if row[1].value == "Услуга 1. Реализация услуг объектов ЖКХ":
                                p2y1FirstRow = row[1].row
                            if row[1].value == "Услуга 2. Реализация услуг объектов спортивно-культурного назначения":
                                p2y2FirstRow = row[1].row
                            if row[1].value == "ИТОГО по Услуге №2":
                                p2y2LastRow = row[1].row
                        p1y1LastRow = p1y2FirstRow - 1
                        p1y2LastRow = p2y1FirstRow - 2
                        p2y1LastRow = p2y2FirstRow - 1
                        # пустая строка для лист 2
                        summRow = []
                        for iCol in range(wbSample.worksheets[1].max_column):
                            summRow.append(None)

                        #  считываем 1 диапазон с данными
                        rows = wsSource.iter_rows(min_row=p1y1FirstRow + 1, max_col=wsSource.max_column, max_row=p1y1LastRow -1, values_only=True)                    
                        for row in rows:
                            if row[1] != None:
                                p1y1.append(list(row))
                                summRow[0] = row[8]
                                #   считываем сумму "Факт прошлого года в разрезе услуг" по Услуге 1
                                try:
                                    summRow[1] = float(summRow[1]) + row[2]
                                except:
                                    summRow[1] = row[2]
                                #   считываем сумму "План текущего года в разрезе услуг по кварталам" по Услуге 1
                                try:
                                    summRow[3] = float(summRow[3]) + row[4]
                                except:
                                    summRow[3] = row[4]
                                try:
                                    summRow[4] = float(summRow[4]) + row[5]
                                except:
                                    summRow[4] = row[5]
                                try:
                                    summRow[5] = float(summRow[5]) + row[6]
                                except:
                                    summRow[5] = row[6]
                                try:
                                    summRow[6] = float(summRow[6]) + row[7]
                                except:
                                    summRow[6] = row[7]
                                #   считываем сумму "Ожидаемое исполнение плана по кварталам" по Услуге 1
                                try:
                                    summRow[8] = float(summRow[8]) + row[19]
                                except:
                                    summRow[8] = row[19]
                                try:
                                    summRow[9] = float(summRow[9]) + row[23]
                                except:
                                    summRow[9] = row[23]
                                try:
                                    summRow[10] = float(summRow[10]) + row[27]
                                except:
                                    summRow[10] = row[27]
                                try:
                                    summRow[11] = float(summRow[11]) + row[31]
                                except:
                                    summRow[11] = row[31]
                        if summRow[0] != None:
                            sp1y1.append(list(summRow))
                        #   1й диапазон с данными - готов
                        
                        #   считываем 2 диапазон с данными
                        # пустая строка для лист 2
                        summRow = []
                        for iCol in range(wbSample.worksheets[1].max_column):
                            summRow.append(None)

                        rows = wsSource.iter_rows(min_row=p1y2FirstRow + 1, max_col=wsSource.max_column, max_row=p1y2LastRow -1, values_only=True)
                        for row in rows:
                            if row[1] != None:
                                p1y2.append(list(row))
                                summRow[0] = row[8]
                                #   считываем сумму "Факт прошлого года в разрезе услуг" по Услуге 2
                                try:
                                    summRow[1] = float(summRow[1]) + row[2]
                                except:
                                    summRow[1] = row[2]
                                #   считываем сумму "План текущего года в разрезе услуг по кварталам" по Услуге 1
                                try:
                                    summRow[3] = float(summRow[3]) + row[4]
                                except:
                                    summRow[3] = row[4]
                                try:
                                    summRow[4] = float(summRow[4]) + row[5]
                                except:
                                    summRow[4] = row[5]
                                try:
                                    summRow[5] = float(summRow[5]) + row[6]
                                except:
                                    summRow[5] = row[6]
                                try:
                                    summRow[6] = float(summRow[6]) + row[7]
                                except:
                                    summRow[6] = row[7]
                                #   считываем сумму "Ожидаемое исполнение плана по кварталам" по Услуге 1
                                try:
                                    summRow[8] = float(summRow[8]) + row[19]
                                except:
                                    summRow[8] = row[19]
                                try:
                                    summRow[9] = float(summRow[9]) + row[23]
                                except:
                                    summRow[9] = row[23]
                                try:
                                    summRow[10] = float(summRow[10]) + row[27]
                                except:
                                    summRow[10] = row[27]
                                try:
                                    summRow[11] = float(summRow[11]) + row[31]
                                except:
                                    summRow[11] = row[31]
                        if summRow[0] != None:
                            sp1y2.append(list(summRow))                            
                        #   2й диапазон с данными - готов
                        #   считываем 3 диапазон с данными. В начало диапазона добавляем строку с итогами по филиалу
                        rows = wsSource.iter_rows(min_row=p2y1FirstRow + 1, max_col=wsSource.max_column, max_row=p2y1LastRow -1, values_only=True)                    
                        # пустая строка для лист 2
                        summRow = []
                        for iCol in range(wbSample.worksheets[1].max_column):
                            summRow.append(None)
                        resultRow = []
                        for iCol in range(wbSample.worksheets[0].max_column):
                            resultRow.append(None)
                        iIndex = 0
                        iIndex = len(p2y1)
                        rowsCount = p2y1LastRow - p2y1FirstRow - 1
                        for row in rows:
                            if row[1] != None:
                                p2y1.append(list(row))
                                resultRow[2] = int(rowsCount)
                                if row[8] != None:
                                    resultRow[8] = row[8]

                                summRow[0] = row[8]
                                #   считываем сумму "Факт прошлого года в разрезе услуг" по Услуге 2
                                try:
                                    summRow[1] = float(summRow[1]) + row[2]
                                except:
                                    summRow[1] = row[2]
                                #   считываем сумму "План текущего года в разрезе услуг по кварталам" по Услуге 1
                                try:
                                    summRow[3] = float(summRow[3]) + row[4]
                                except:
                                    summRow[3] = row[4]
                                try:
                                    summRow[4] = float(summRow[4]) + row[5]
                                except:
                                    summRow[4] = row[5]
                                try:
                                    summRow[5] = float(summRow[5]) + row[6]
                                except:
                                    summRow[5] = row[6]
                                try:
                                    summRow[6] = float(summRow[6]) + row[7]
                                except:
                                    summRow[6] = row[7]
                                #   считываем сумму "Ожидаемое исполнение плана по кварталам" по Услуге 1
                                try:
                                    summRow[8] = float(summRow[8]) + row[19]
                                except:
                                    summRow[8] = row[19]
                                try:
                                    summRow[9] = float(summRow[9]) + row[23]
                                except:
                                    summRow[9] = row[23]
                                try:
                                    summRow[10] = float(summRow[10]) + row[27]
                                except:
                                    summRow[10] = row[27]
                                try:
                                    summRow[11] = float(summRow[11]) + row[31]
                                except:
                                    summRow[11] = row[31]
                        if summRow[0] != None:
                            sp2y1.append(list(summRow))                            

                        if iIndex < len(p2y1):
                            p2y1.insert(iIndex, resultRow)
                        #   3й диапазон с данными - готов
                        #   считываем 4 диапазон с данными. В начало диапазона добавляем строку с итогами по филиалу
                        rows = wsSource.iter_rows(min_row=p2y2FirstRow + 1, max_col=wsSource.max_column, max_row=p2y2LastRow -1, values_only=True)
                        summRow = []
                        for iCol in range(wbSample.worksheets[1].max_column):
                            summRow.append(None)
                        resultRow = []
                        for iCol in range(wbSample.worksheets[0].max_column):
                            resultRow.append(None)
                        iIndex = len(p2y2)
                        rowsCount = p2y2LastRow - p2y2FirstRow - 1
                        for row in rows:
                            if row[1] != None:
                                p2y2.append(list(row))
                                resultRow[2] = int(rowsCount)
                                if row[8] != None:
                                    resultRow[8] = row[8]
                                summRow[0] = row[8]
                                #   считываем сумму "Факт прошлого года в разрезе услуг" по Услуге 2
                                try:
                                    summRow[1] = float(summRow[1]) + row[2]
                                except:
                                    summRow[1] = row[2]
                                #   считываем сумму "План текущего года в разрезе услуг по кварталам" по Услуге 1
                                try:
                                    summRow[3] = float(summRow[3]) + row[4]
                                except:
                                    summRow[3] = row[4]
                                try:
                                    summRow[4] = float(summRow[4]) + row[5]
                                except:
                                    summRow[4] = row[5]
                                try:
                                    summRow[5] = float(summRow[5]) + row[6]
                                except:
                                    summRow[5] = row[6]
                                try:
                                    summRow[6] = float(summRow[6]) + row[7]
                                except:
                                    summRow[6] = row[7]
                                #   считываем сумму "Ожидаемое исполнение плана по кварталам" по Услуге 1
                                try:
                                    summRow[8] = float(summRow[8]) + row[19]
                                except:
                                    summRow[8] = row[19]
                                try:
                                    summRow[9] = float(summRow[9]) + row[23]
                                except:
                                    summRow[9] = row[23]
                                try:
                                    summRow[10] = float(summRow[10]) + row[27]
                                except:
                                    summRow[10] = row[27]
                                try:
                                    summRow[11] = float(summRow[11]) + row[31]
                                except:
                                    summRow[11] = row[31]
                        if summRow[0] != None:
                            sp2y2.append(list(summRow))                            
                        if iIndex < len(p2y2):
                            p2y2.insert(iIndex, resultRow)
                        wbSource.close()
                    except: # ошибка обработки файла iSourceFileName
                        py_logger.error(f"Ошибка обработки файла {iSourceFileName}", exc_info=True)
            ###########################################
                # формируем 1 блок данных в свод лист 2
                if len(sp1y1) > 0:
                    sp1y1FirstRow = wsReport02.max_row + 1
                    sp1y1LastRow = wsReport02.max_row + len(sp1y1)
                    for iRow in sp1y1:                
                        wsReport02.append(iRow)

                    for iRow in range(sp1y1FirstRow, sp1y1LastRow+1):
                        for iColumn in range(1, len(sp1y1[0])+1):
                            copy_cell(wbSample.worksheets[1], 8, iColumn,
                                    wsReport02, iRow, iColumn, copy_style_only=True)
                    # копируем строки с 10 по 11 из образца в конец свода лист 2
                    for iRow in wbSample.worksheets[1].iter_rows(min_row=10, max_col=wbSample.worksheets[1].max_column, max_row=11):
                        lRow = wsReport02.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[1], iCell.row, iCell.column,
                                    wsReport02, lRow, iCell.column)                
                            # корректировка формул лист 2
                            if iCell.data_type == 'f':
                                formulaCorrection = str(iCell.value).split(':')
                                if len(formulaCorrection) > 1:
                                    formulaCorrection[0] = str(formulaCorrection[0]).replace('8', str(sp1y1FirstRow))
                                    formulaCorrection[1] = str(formulaCorrection[1]).replace('9', str(sp1y1LastRow))
                                    wsReport02.cell(lRow, iCell.column).value = ':'.join(formulaCorrection)
                else:
                    # копируем строки с 8 по 11 из образца в конец свода  лист 2
                    for iRow in wbSample.worksheets[1].iter_rows(min_row=8, max_col=wbSample.worksheets[1].max_column, max_row=11):
                        lRow = wsReport02.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[1], iCell.row, iCell.column,
                                    wsReport02, lRow, iCell.column)
                
                # формируем 1 блок данных в свод лист 1
                if len(p1y1) > 0:
                    p1y1FirstRow = wsReport01.max_row + 1
                    p1y1LastRow = wsReport01.max_row + len(p1y1)
                    iCount = 0
                    for iRow in p1y1:                
                        iCount += 1
                        iRow[0] = '1.' + str(iCount)
                        wsReport01.append(iRow)

                    for iRow in range(p1y1FirstRow, p1y1LastRow+1):
                        for iColumn in range(1, len(p1y1[0])+1):
                            copy_cell(wbSample.worksheets[0], 8, iColumn,
                                    wsReport01, iRow, iColumn, copy_style_only=True)
                    # копируем строки с 10 по 11 из образца в конец свода лист 1
                    for iRow in wbSample.worksheets[0].iter_rows(min_row=10, max_col=wbSample.worksheets[0].max_column, max_row=11):
                        lRow = wsReport01.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[0], iCell.row, iCell.column,
                                    wsReport01, lRow, iCell.column)
                            # корректировка формул. 
                            if iCell.data_type == 'f':
                                formulaCorrection = str(iCell.value).split(':')
                                if len(formulaCorrection) > 1:
                                    formulaCorrection[0] = str(formulaCorrection[0]).replace('8', str(p1y1FirstRow))
                                    formulaCorrection[1] = str(formulaCorrection[1]).replace('9', str(p1y1LastRow))
                                    wsReport01.cell(lRow, iCell.column).value = ':'.join(formulaCorrection)
                else:
                    p1y1FirstRow = wsReport01.max_row + 1
                    p1y1LastRow = wsReport01.max_row + 2
                    # копируем строки с 8 по 11 из образца в конец свода лист 1
                    for iRow in wbSample.worksheets[0].iter_rows(min_row=8, max_col=wbSample.worksheets[0].max_column, max_row=11):
                        lRow = wsReport01.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[0], iCell.row, iCell.column,
                                    wsReport01, lRow, iCell.column)
                
                # формируем 2 блок данных в свод лист 2
                if len(sp1y2) > 0:
                    sp1y2FirstRow = wsReport02.max_row + 1
                    sp1y2LastRow = wsReport02.max_row + len(sp1y2)
                    
                    for iRow in sp1y2:                
                        wsReport02.append(iRow)
                    for iRow in range(sp1y2FirstRow, sp1y2LastRow+1):
                        for iColumn in range(1, len(sp1y2[0])+1):
                            copy_cell(wbSample.worksheets[1], 12, iColumn,
                                    wsReport02, iRow, iColumn, copy_style_only=True)
                    # копируем строки с 14 по 16 из образца в конец свода
                    for iRow in wbSample.worksheets[1].iter_rows(min_row=14, max_col=wbSample.worksheets[1].max_column, max_row=16):
                        lRow = wsReport02.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[1], iCell.row, iCell.column,
                                    wsReport02, lRow, iCell.column)
                            if iCell.data_type == 'f':
                                formulaCorrection = str(iCell.value).split(':')
                                if len(formulaCorrection) > 1:
                                    formulaCorrection[0] = str(formulaCorrection[0]).replace('12', str(sp1y2FirstRow))
                                    formulaCorrection[1] = str(formulaCorrection[1]).replace('13', str(sp1y2LastRow))
                                    wsReport02.cell(lRow, iCell.column).value = ':'.join(formulaCorrection)
                else:
                    sp1y2FirstRow = wsReport02.max_row + 1
                    sp1y2LastRow = wsReport02.max_row + 2
                    # копируем строки с 12 по 16 из образца в конец свода
                    for iRow in wbSample.worksheets[1].iter_rows(min_row=12, max_col=wbSample.worksheets[1].max_column, max_row=16):
                        lRow = wsReport02.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[1], iCell.row, iCell.column,
                                    wsReport02, lRow, iCell.column)
                            if iCell.data_type == 'f':
                                formulaCorrection = str(iCell.value).split(':')
                                if len(formulaCorrection) > 1:
                                    formulaCorrection[0] = str(formulaCorrection[0]).replace('12', str(sp1y2FirstRow))
                                    formulaCorrection[1] = str(formulaCorrection[1]).replace('13', str(sp1y2LastRow))
                                    wsReport02.cell(lRow, iCell.column).value = ':'.join(formulaCorrection)

                # формируем 2 блок данных в свод лист 1
                if len(p1y2) > 0:
                    p1y2FirstRow = wsReport01.max_row + 1
                    p1y2LastRow = wsReport01.max_row + len(p1y2)
                    iCount = 0
                    for iRow in p1y2:
                        iCount += 1
                        iRow[0] = '2.' + str(iCount)
                        wsReport01.append(iRow)

                    for iRow in range(p1y2FirstRow, p1y2LastRow+1):
                        for iColumn in range(1, len(p1y2[0])+1):
                            copy_cell(wbSample.worksheets[0], 12, iColumn,
                                    wsReport01, iRow, iColumn, copy_style_only=True)
                    # копируем строки с 14 по 16 из образца в конец свода
                    for iRow in wbSample.worksheets[0].iter_rows(min_row=14, max_col=wbSample.worksheets[0].max_column, max_row=16):
                        lRow = wsReport01.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[0], iCell.row, iCell.column,
                                    wsReport01, lRow, iCell.column)
                            if iCell.data_type == 'f':
                                formulaCorrection = str(iCell.value).split(':')
                                formulaCorrection[0] = str(formulaCorrection[0]).replace('12', str(p1y2FirstRow))
                                formulaCorrection[1] = str(formulaCorrection[1]).replace('13', str(p1y2LastRow))
                                wsReport01.cell(lRow, iCell.column).value = ':'.join(formulaCorrection)
                else:
                    p1y2FirstRow = wsReport01.max_row + 1
                    p1y2LastRow = wsReport01.max_row + 2
                    # копируем строки с 12 по 16 из образца в конец свода
                    for iRow in wbSample.worksheets[0].iter_rows(min_row=12, max_col=wbSample.worksheets[0].max_column, max_row=16):
                        lRow = wsReport01.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[0], iCell.row, iCell.column,
                                    wsReport01, lRow, iCell.column)
                            if iCell.data_type == 'f':
                                formulaCorrection = str(iCell.value).split(':')
                                if len(formulaCorrection) > 1:
                                    formulaCorrection[0] = str(formulaCorrection[0]).replace('12', str(p1y2FirstRow))
                                    formulaCorrection[1] = str(formulaCorrection[1]).replace('13', str(p1y2LastRow))
                                    wsReport01.cell(lRow, iCell.column).value = ':'.join(formulaCorrection)
                # формируем 3 блок данных в свод лист 2
                if len(sp2y1) > 0:
                    sp2y1FirstRow = wsReport02.max_row + 1
                    sp2y1LastRow = wsReport02.max_row + len(sp2y1)
                    
                    for iRow in sp2y1:                
                        wsReport02.append(iRow)
                    for iRow in range(sp2y1FirstRow, sp2y1LastRow+1):
                        for iColumn in range(1, len(sp2y1[0])+1):
                            copy_cell(wbSample.worksheets[1], 17, iColumn,
                                    wsReport02, iRow, iColumn, copy_style_only=True)
                    # копируем строки с 19 по 20 из образца в конец свода
                    for iRow in wbSample.worksheets[1].iter_rows(min_row=19, max_col=wbSample.worksheets[1].max_column, max_row=20):
                        lRow = wsReport02.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[1], iCell.row, iCell.column,
                                    wsReport02, lRow, iCell.column)
                            if iCell.data_type == 'f':
                                formulaCorrection = str(iCell.value).split(':')
                                if len(formulaCorrection) > 1:
                                    formulaCorrection[0] = str(formulaCorrection[0]).replace('17', str(sp2y1FirstRow))
                                    formulaCorrection[1] = str(formulaCorrection[1]).replace('18', str(sp2y1LastRow))
                                    wsReport02.cell(lRow, iCell.column).value = ':'.join(formulaCorrection)
                else:
                    sp2y1FirstRow = wsReport02.max_row + 1
                    sp2y1LastRow = wsReport02.max_row + 2
                    # копируем строки с 17 по 20 из образца в конец свода
                    for iRow in wbSample.worksheets[1].iter_rows(min_row=17, max_col=wbSample.worksheets[1].max_column, max_row=20):
                        lRow = wsReport02.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[1], iCell.row, iCell.column,
                                    wsReport02, lRow, iCell.column)
                            if iCell.data_type == 'f':
                                formulaCorrection = str(iCell.value).split(':')
                                if len(formulaCorrection) > 1:
                                    formulaCorrection[0] = str(formulaCorrection[0]).replace('17', str(sp2y1FirstRow))
                                    formulaCorrection[1] = str(formulaCorrection[1]).replace('18', str(sp2y1LastRow))
                                    wsReport02.cell(lRow, iCell.column).value = ':'.join(formulaCorrection)
                # формируем 3 блок данных в свод лист 1
                if len(p2y1) > 0:
                    p2y1FirstRow = wsReport01.max_row + 1
                    p2y1LastRow = wsReport01.max_row + len(p2y1)
                    
                    iCount = 0
                    for iRow in p2y1:
                        if iRow[1] != None:
                            iCount += 1
                            iRow[0] = '1.' + str(iCount)
                        wsReport01.append(iRow)
                    
                    for iRow in range(p2y1FirstRow, p2y1LastRow+1):
                        for iColumn in range(1, len(p2y1[0])+1):
                            copy_cell(wbSample.worksheets[0], 17, iColumn,
                                    wsReport01, iRow, iColumn, copy_style_only=True)
                        if wsReport01.cell(iRow, 2).value == None:
                            rowsCount = int(wsReport01.cell(iRow, 3).value)
                            for iColumn in range(3, 9):
                                wsReport01.cell(iRow, iColumn).value = "=SUM(" + wsReport01.cell(iRow, iColumn).column_letter + str(iRow + 1) + ":" + wsReport01.cell(iRow, iColumn).column_letter + str(iRow + rowsCount) + ")"
                            for iColumn in range(15, 35):
                                wsReport01.cell(iRow, iColumn).value = "=SUM(" + wsReport01.cell(iRow, iColumn).column_letter + str(iRow + 1) + ":" + wsReport01.cell(iRow, iColumn).column_letter + str(iRow + rowsCount) + ")"                    
                            for iColumn in range(1, len(p2y1[0])+1):
                                wsReport01.cell(iRow, iColumn).font = copy(wbSample.worksheets[0].cell(5, iColumn).font)
                                wsReport01.cell(iRow, iColumn).fill = copy(wbSample.worksheets[0].cell(5, iColumn).fill)

                    # копируем строки с 19 по 20 из образца в конец свода
                    for iRow in wbSample.worksheets[0].iter_rows(min_row=19, max_col=wbSample.worksheets[0].max_column, max_row=20):
                        lRow = wsReport01.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[0], iCell.row, iCell.column,
                                    wsReport01, lRow, iCell.column)
                            if iCell.data_type == 'f':
                                formulaCorrection = str(iCell.value).split(':')
                                if len(formulaCorrection) > 1:
                                    formulaCorrection[0] = str(formulaCorrection[0]).replace('17', str(p2y1FirstRow))
                                    formulaCorrection[1] = str(formulaCorrection[1]).replace('18', str(p2y1LastRow))
                                    wsReport01.cell(lRow, iCell.column).value = ':'.join(formulaCorrection)
                else:
                    p2y1FirstRow = wsReport01.max_row + 1
                    p2y1LastRow = wsReport01.max_row + 2
                    # копируем строки с 17 по 20 из образца в конец свода
                    for iRow in wbSample.worksheets[0].iter_rows(min_row=17, max_col=wbSample.worksheets[0].max_column, max_row=20):
                        lRow = wsReport01.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[0], iCell.row, iCell.column,
                                    wsReport01, lRow, iCell.column)
                            if iCell.data_type == 'f':
                                formulaCorrection = str(iCell.value).split(':')
                                if len(formulaCorrection) > 1:
                                    formulaCorrection[0] = str(formulaCorrection[0]).replace('17', str(p2y1FirstRow))
                                    formulaCorrection[1] = str(formulaCorrection[1]).replace('18', str(p2y1LastRow))
                                    wsReport01.cell(lRow, iCell.column).value = ':'.join(formulaCorrection)
                # формируем 4 блок данных в свод лист 2
                if len(sp2y2) > 0:
                    sp2y2FirstRow = wsReport02.max_row + 1
                    sp2y2LastRow = wsReport02.max_row + len(sp2y2)
                    for iRow in sp2y2:                
                        wsReport02.append(iRow)
                    for iRow in range(sp2y2FirstRow, sp2y2LastRow+1):
                        for iColumn in range(1, len(sp2y2[0])+1):
                            copy_cell(wbSample.worksheets[1], 21, iColumn,
                                    wsReport02, iRow, iColumn, copy_style_only=True)
                    # копируем строку 23 из образца в конец свода
                    for iRow in wbSample.worksheets[1].iter_rows(min_row=23, max_col=wbSample.worksheets[1].max_column, max_row=23):
                        lRow = wsReport02.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[1], iCell.row, iCell.column,
                                    wsReport02, lRow, iCell.column)
                            if iCell.data_type == 'f':
                                formulaCorrection = str(iCell.value).split(':')
                                if len(formulaCorrection) > 1:
                                    formulaCorrection[0] = str(formulaCorrection[0]).replace('21', str(sp2y2FirstRow))
                                    formulaCorrection[1] = str(formulaCorrection[1]).replace('22', str(sp2y2LastRow))
                                    wsReport02.cell(lRow, iCell.column).value = ':'.join(formulaCorrection)
                    # копируем строку 24 из образца в конец свода
                    for iRow in wbSample.worksheets[1].iter_rows(min_row=24, max_col=wbSample.worksheets[1].max_column, max_row=24):
                        lRow = wsReport02.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[1], iCell.row, iCell.column,
                                    wsReport02, lRow, iCell.column)
                            if iCell.data_type == 'f':
                                formulaCorrection = str(iCell.value).split('+')
                                if len(formulaCorrection) > 3:
                                    formulaCorrection[3] = formulaCorrection[3].replace('10', str(sp1y1LastRow + 1))
                                    formulaCorrection[2] = formulaCorrection[2].replace('14', str(sp1y2LastRow + 1))
                                    formulaCorrection[1] = formulaCorrection[1].replace('19', str(sp2y1LastRow + 1))
                                    formulaCorrection[0] = formulaCorrection[0].replace('23', str(sp2y2LastRow + 1))
                                    wsReport02.cell(lRow, iCell.column).value = '+'.join(formulaCorrection)
                    
                # формируем 4 блок данных в свод лист 1
                if len(p2y2) > 0:
                    p2y2FirstRow = wsReport01.max_row + 1
                    p2y2LastRow = wsReport01.max_row + len(p2y2)
                    iCount = 0
                    for iRow in p2y2:
                        if iRow[1] != None:
                            iCount += 1
                            iRow[0] = '2.' + str(iCount)
                        wsReport01.append(iRow)

                    for iRow in range(p2y2FirstRow, p2y2LastRow+1):
                        for iColumn in range(1, len(p2y2[0])+1):
                            copy_cell(wbSample.worksheets[0], 21, iColumn,
                                    wsReport01, iRow, iColumn, copy_style_only=True)
                        if wsReport01.cell(iRow, 2).value == None:
                            rowsCount = int(wsReport01.cell(iRow, 3).value)
                            for iColumn in range(3, 9):
                                wsReport01.cell(iRow, iColumn).value = "=SUM(" + wsReport01.cell(iRow, iColumn).column_letter + str(iRow + 1) + ":" + wsReport01.cell(iRow, iColumn).column_letter + str(iRow + rowsCount) + ")"
                            for iColumn in range(15, 35):
                                wsReport01.cell(iRow, iColumn).value = "=SUM(" + wsReport01.cell(iRow, iColumn).column_letter + str(iRow + 1) + ":" + wsReport01.cell(iRow, iColumn).column_letter + str(iRow + rowsCount) + ")"
                            for iColumn in range(1, len(p2y2[0])+1):
                                wsReport01.cell(iRow, iColumn).font = copy(wbSample.worksheets[0].cell(5, iColumn).font)
                                wsReport01.cell(iRow, iColumn).fill = copy(wbSample.worksheets[0].cell(5, iColumn).fill)

                    # копируем строку 23 из образца в конец свода
                    for iRow in wbSample.worksheets[0].iter_rows(min_row=23, max_col=wbSample.worksheets[0].max_column, max_row=23):
                        lRow = wsReport01.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[0], iCell.row, iCell.column,
                                    wsReport01, lRow, iCell.column)
                            if iCell.data_type == 'f':
                                formulaCorrection = str(iCell.value).split(':')
                                if len(formulaCorrection) > 1:
                                    formulaCorrection[0] = str(formulaCorrection[0]).replace('21', str(p2y2FirstRow))
                                    formulaCorrection[1] = str(formulaCorrection[1]).replace('22', str(p2y2LastRow))
                                    wsReport01.cell(lRow, iCell.column).value = ':'.join(formulaCorrection)
                    # копируем строку 24 из образца в конец свода
                    for iRow in wbSample.worksheets[0].iter_rows(min_row=24, max_col=wbSample.worksheets[0].max_column, max_row=24):
                        lRow = wsReport01.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[0], iCell.row, iCell.column,
                                    wsReport01, lRow, iCell.column)
                            if iCell.data_type == 'f':
                                formulaCorrection = str(iCell.value).split('+')
                                formulaCorrection[3] = formulaCorrection[3].replace('10', str(p1y1LastRow + 1))
                                formulaCorrection[2] = formulaCorrection[2].replace('14', str(p1y2LastRow + 1))
                                formulaCorrection[1] = formulaCorrection[1].replace('19', str(p2y1LastRow + 1))
                                formulaCorrection[0] = formulaCorrection[0].replace('23', str(p2y2LastRow + 1))
                                wsReport01.cell(lRow, iCell.column).value = '+'.join(formulaCorrection)
                    # копируем строки с 25 по 32 из образца в конец свода
                    for iRow in wbSample.worksheets[0].iter_rows(min_row=25, max_col=wbSample.worksheets[0].max_column, max_row=32):
                        lRow = wsReport01.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[0], iCell.row, iCell.column,
                                    wsReport01, lRow, iCell.column)
                else:
                    p2y2FirstRow = wsReport01.max_row + 1
                    p2y2LastRow = wsReport01.max_row + 2
                    # копируем строки с 21 по 22 из образца в конец свода
                    for iRow in wbSample.worksheets[0].iter_rows(min_row=21, max_col=wbSample.worksheets[0].max_column, max_row=22):
                        lRow = wsReport01.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[0], iCell.row, iCell.column,
                                    wsReport01, lRow, iCell.column)
                    # копируем строку 23 из образца в конец свода
                    for iRow in wbSample.worksheets[0].iter_rows(min_row=23, max_col=wbSample.worksheets[0].max_column, max_row=23):
                        lRow = wsReport01.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[0], iCell.row, iCell.column,
                                    wsReport01, lRow, iCell.column)
                            if iCell.data_type == 'f':
                                formulaCorrection = str(iCell.value).split(':')
                                if len(formulaCorrection) > 1:
                                    formulaCorrection[0] = str(formulaCorrection[0]).replace('21', str(p2y2FirstRow))
                                    formulaCorrection[1] = str(formulaCorrection[1]).replace('22', str(p2y2LastRow))
                                    wsReport01.cell(lRow, iCell.column).value = ':'.join(formulaCorrection)
                    # копируем строку 24 из образца в конец свода
                    for iRow in wbSample.worksheets[0].iter_rows(min_row=24, max_col=wbSample.worksheets[0].max_column, max_row=24):
                        lRow = wsReport01.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[0], iCell.row, iCell.column,
                                    wsReport01, lRow, iCell.column)
                            if iCell.data_type == 'f':
                                formulaCorrection = str(iCell.value).split('+')
                                formulaCorrection[3] = formulaCorrection[3].replace('10', str(p1y1LastRow + 1))
                                formulaCorrection[2] = formulaCorrection[2].replace('14', str(p1y2LastRow + 1))
                                formulaCorrection[1] = formulaCorrection[1].replace('19', str(p2y1LastRow + 1))
                                formulaCorrection[0] = formulaCorrection[0].replace('23', str(p2y2LastRow + 1))
                                wsReport01.cell(lRow, iCell.column).value = '+'.join(formulaCorrection)
                    # копируем строки с 25 по 32 из образца в конец свода
                    for iRow in wbSample.worksheets[0].iter_rows(min_row=25, max_col=wbSample.worksheets[0].max_column, max_row=32):
                        lRow = wsReport01.max_row+1
                        for iCell in iRow:
                            copy_cell(wbSample.worksheets[0], iCell.row, iCell.column,
                                    wsReport01, lRow, iCell.column)        
                # высота строк с 7 - авто
                for iRow in range(7, wsReport01.max_row):
                    wsReport01.row_dimensions[iRow].height = None
                for iRow in range(7, wsReport01.max_row):
                    wsReport02.row_dimensions[iRow].height = None
                # группировка столбцов
                wsReport01.column_dimensions.group('E','H', hidden=True)
                wsReport01.column_dimensions.group('Q','S', hidden=True)
                wsReport01.column_dimensions.group('U','W', hidden=True)
                wsReport01.column_dimensions.group('Y','AA', hidden=True)
                wsReport01.column_dimensions.group('AC','AE', hidden=True)

                wbReport.save(wbReportFileName)  # сохраняем результат
                wbReport.close()
                wbSample.close()
                py_logger.info(f"Обработка файлов выполнена успешно, результат обработки {wbReportFileName}")    

        FilialsDict = dict()
        FilialsDict = {
            2410: "АУП",
            2418: "Белоярское УАВР",
            2422: "Белоярское УТТиСТ",
            2430: "Бобровское ЛПУМГ",
            2431: "Верхнеказымское ЛПУМГ",
            2432: "Ивдельское ЛПУМГ",
            2424: "ИТЦ",
            2433: "Казымское ЛПУМГ",
            2434: "Карпинское ЛПУМГ",
            2435: "Комсомольское ЛПУМГ",
            2436: "Краснотурьинское ЛПУМГ",
            2471: "КСК НОРД",
            2437: "Лонг-Юганское ЛПУМГ",
            2438: "Надымское ЛПУМГ",
            2416: "Надымское УАВР",
            2439: "Нижнетуринское ЛПУМГ",
            2440: "Ново-Уренгойское ЛПУМГ",
            2420: "Надымское УТТиСТ",
            2441: "Ныдинское ЛПУМГ",
            2442: "Октябрьское ЛПУМГ",
            2443: "Пангодинское ЛПУМГ",
            2444: "Пелымское ЛПУМГ",
            2445: "Перегребненское ЛПУМГ",
            2446: "Правохеттинское ЛПУМГ",
            2447: "Приозерное ЛПУМГ",
            2448: "Пунгинское ЛПУМГ",
            2470: "Санаторий-профилакторий",
            2449: "Сорумское ЛПУМГ",
            2450: "Сосновское ЛПУМГ",
            2451: "Сосьвинское ЛПУМГ",
            2452: "Таёжное ЛПУМГ",
            2414: "УОВОФ",
            2425: "УПЦ",
            2453: "Уральское ЛПУМГ",
            2427: "Управление связи",
            2426: "УЭЗиС",
            2417: "Югорское УАВР",
            2421: "Югорское УТТиСТ",
            2411: "Югорское УМТСиК",
            2454: "Ягельное ЛПУМГ",
            2455: "Ямбургское ЛПУМГ"}

        files = os.listdir(pluginInput)

        wbSources = list()
        wbSampleFN = ""
        wbReportFileName = ""
        wbSources.clear()
        for iFile in files:
            file_name, file_extension = os.path.splitext(iFile)
            if file_extension == ".xlsx":
                if file_name.find("ОСР_Ожидаемое доходы(свод)") != -1:
                    wbSampleFN = pluginInput + "/" + iFile
                    wbReportFileName = pluginOutput + "/" + "Ожидаемое доходы(свод).xlsx"
                for icode in FilialsDict.keys():
                    if file_name.find(str(icode)) != -1:
                        wbSources.append(pluginInput + "/" + iFile)

        
        try:
            py_logger.info(f"Запускаем обработку файлов {wbSources}, образец {wbSampleFN}, результат обработки выводим в файл {wbReportFileName}")
            makeReport(wbSources, wbSampleFN, wbReportFileName)
        except:
            py_logger.error(f"Ошибка при обработке файлов", exc_info=True)



